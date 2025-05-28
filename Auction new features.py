import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re

class FileSelectPage(tk.Frame):
    def __init__(self, master, on_file_loaded):
        super().__init__(master)
        self.on_file_loaded = on_file_loaded
        tk.Label(self, text="Select Excel/CSV file", font=("Arial", 16)).pack(pady=20)
        tk.Button(self, text="Browse", command=self.browse_file, font=("Arial", 14)).pack(pady=10)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xls;*.xlsx")]
        )
        if file_path:
            try:
                df = pd.read_excel(file_path, sheet_name=None)
                sheet = list(df.values())[0]
                # Find the split (blank row)
                blank_idx = sheet[sheet.isnull().all(axis=1)].index[0]
                teams_df = sheet.iloc[:blank_idx, :2].copy()
                players_df = sheet.iloc[blank_idx+2:, :2].copy()
                teams_df.columns = ["Team name", "Team starting money"]
                players_df.columns = ["Player name", "Bid value"]
                # Check for missing values
                if teams_df.isnull().any().any() or players_df.isnull().any().any():
                    messagebox.showerror("Error", "File contains missing values. Please fill all cells.")
                    return
                self.on_file_loaded(teams_df, players_df)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load file: {e}")

class AuctionApp(tk.Frame):
    def __init__(self, master, teams_df, players_df, auction_name="AuctionSession"):
        super().__init__(master)
        self.master = master
        self.pack(fill="both", expand=True)
        self.auction_name = auction_name
        self.excel_filename = f"{auction_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.setup_excel()

        # Parse teams and players from dataframes
        self.teams = list(teams_df["Team name"])
        self.manager_ids = {team: i+1 for i, team in enumerate(self.teams)}
        self.team_money = {row["Team name"]: int(row["Team starting money"]) for _, row in teams_df.iterrows()}
        self.team_inventory = {team: {} for team in self.teams}
        self.items = [(row["Player name"], int(row["Bid value"])) for _, row in players_df.iterrows()]
        self.player_ids = {name: 101+i for i, (name, _) in enumerate(self.items)}
        self.write_header_to_excel()

        # --- GUI Layout ---
        self.money_labels = {}
        self.inventory_heading_labels = {}
        self.inventory_list_labels = {}
        self.bid_history = []
        self.current_item = None
        self.current_bid = None
        self.highest_bidder = None
        self.bidding_enabled = 0

        # Current item label
        self.current_item_label = tk.Label(
            self,
            text="Current Item: ",
            font=('Arial', 14, 'bold'),
            width=50,
            anchor="w"
        )
        self.current_item_label.grid(row=0, column=0, columnspan=len(self.teams), pady=10, padx=5, sticky="w")

        # Bidding status label
        self.bid_status_label = tk.Label(
            self,
            text="Bidding Status: ",
            font=('Arial', 14)
        )
        self.bid_status_label.grid(row=1, column=0, columnspan=len(self.teams), pady=10, padx=5, sticky="w")

        # Team columns
        for i, team in enumerate(self.teams):
            # Money label
            money_label = tk.Label(
                self,
                text=f"{team} \nMoney: ₹{self.team_money[team]}",
                font=('Arial', 12, 'bold'),
                width=20,
                anchor="n"
            )
            money_label.grid(row=2, column=i, pady=3, padx=3, sticky="nsew")
            self.money_labels[team] = money_label

            # Create a frame for each team's inventory (heading + list)
            inventory_frame = tk.Frame(self)
            inventory_frame.grid(row=3, column=i, rowspan=2, pady=3, padx=3, sticky="nsew")

            # Inventory heading label (inside frame)
            inventory_heading_label = tk.Label(
                inventory_frame,
                text=f"{team} \nInventory:",
                font=('Arial', 12, 'bold'),
                width=20,
                anchor="n"
            )
            inventory_heading_label.pack(side="top", fill="x")

            # Inventory list label (inside frame)
            inventory_list_label = tk.Label(
                inventory_frame,
                text="",
                font=('Arial', 12),
                width=20,
                wraplength=200,
                justify="center",
                anchor="n"
            )
            inventory_list_label.pack(side="top", fill="both", expand=True)

            self.inventory_heading_labels[team] = inventory_heading_label
            self.inventory_list_labels[team] = inventory_list_label

            # Bid button
            bid_button = tk.Button(
                self,
                text=f"Bid",
                font=('Arial', 12),
                command=lambda team=team: self.place_bid(team),
                width=12,
                height=2,
                bg="lightblue"
            )
            bid_button.grid(row=10, column=i, pady=5, padx=5, sticky="s")

        # Undo Button
        self.undo_button = tk.Button(
            self,
            text="Undo",
            width=12,
            height=2,
            bg="orange",
            font=('Arial', 12),
            command=self.undo_last_bid
        )
        self.undo_button.grid(row=15, column=3, columnspan=1, pady=10, padx=5, sticky="nsew")

        # End Bidding Button
        end_bidding_button = tk.Button(
            self,
            text="End Bidding Round",
            height=2,
            bg="red",
            font=('Arial', 12),
            command=self.end_bidding_round
        )
        end_bidding_button.grid(row=15, column=1, columnspan=2, pady=10, padx=5, sticky="nsew")

        # Player list (scrolled text) in the rightmost column
        player_list_col = len(self.teams)
        self.item_list_text = scrolledtext.ScrolledText(
            self,
            width=40,
            height=30,
            font=('Arial', 12),
            wrap=tk.WORD
        )
        self.item_list_text.grid(row=2, column=player_list_col, rowspan=8, padx=5, pady=5, sticky="nsew")
        for item, price in self.items:
            item_button = tk.Button(
                self,
                text=f"{item} (Starting Price: ₹{price})",
                font=('Arial', 12),
                command=lambda item=(item, price): self.select_item(item)
            )
            item_button.grid(sticky="w")
            self.item_list_text.window_create("end", window=item_button)
            self.item_list_text.insert("end", "\n")

        # --- Grid configuration for resizing ---
        for i in range(16):
            self.rowconfigure(i, weight=1)
        for i in range(len(self.teams)):
            self.columnconfigure(i, weight=1)
        self.columnconfigure(player_list_col, weight=0, minsize=320)

    # --- Excel Integration Methods ---

    def setup_excel(self):
        """Create or open an Excel workbook and add a new session sheet for this session."""
        if os.path.exists(self.excel_filename):
            self.wb = load_workbook(self.excel_filename)
            # Find the next available session number
            session_base = "Session"
            session_nums = [int(re.search(rf"{session_base}_(\d+)", ws).group(1))
                            for ws in self.wb.sheetnames if re.match(rf"{session_base}_\d+", ws)]
            next_num = max(session_nums + [1]) + 1 if session_nums else 2
            session_name = f"{session_base}_{next_num}"
        else:
            self.wb = Workbook()
            session_name = "Session_1"
            # Remove default sheet if present
            if "Sheet" in self.wb.sheetnames:
                std = self.wb["Sheet"]
                self.wb.remove(std)
        self.session_ws = self.wb.create_sheet(session_name)
        self.wb.save(self.excel_filename)

    def write_header_to_excel(self):
        """Write auction header info, player list, and prepare StateLog in the same sheet."""
        ws = self.session_ws
        ws["A1"] = "Auction Name"
        ws["B1"] = self.auction_name
        ws["A2"] = "Date"
        ws["B2"] = datetime.now().strftime("%Y-%m-%d")
        ws["A3"] = "Time"
        ws["B3"] = datetime.now().strftime("%H:%M:%S")
        ws["A4"] = "Total Players"
        ws["B4"] = len(self.items)

        # Player databse header
        ws["A6"] = "PlayerID"
        ws["B6"] = "Player Name"
        ws["C6"] = "Base Bid Value"

        # Team info header (starting from column E)
        ws["E6"] = "Team Name"
        ws["F6"] = "Team ID"
        ws["G6"] = "Starting Money"
        ws["H6"] = "End Money"
        # Fill player info
        for idx, (name, base) in enumerate(self.items):
            ws[f"A{7+idx}"] = self.player_ids[name]
            ws[f"B{7+idx}"] = name
            ws[f"C{7+idx}"] = base

        # Fill team info
        for idx, team in enumerate(self.teams):
            ws[f"E{7+idx}"] = team
            ws[f"F{7+idx}"] = self.manager_ids[team]
            ws[f"G{7+idx}"] = self.team_money[team]  # Starting money
            ws[f"H{7+idx}"] = self.team_money[team]  # End money (will be updated at end)
            
        # Leave a blank row, then start StateLog
        self.statelog_start_row = 8 + len(self.items)
        ws[f"A{self.statelog_start_row}"] = "Timestamp"
        ws[f"B{self.statelog_start_row}"] = "Event"
        ws[f"C{self.statelog_start_row}"] = "ManagerID"
        ws[f"D{self.statelog_start_row}"] = "ManagerName"
        ws[f"E{self.statelog_start_row}"] = "PlayerID"
        ws[f"F{self.statelog_start_row}"] = "PlayerName"
        ws[f"G{self.statelog_start_row}"] = "BaseBid"
        ws[f"H{self.statelog_start_row}"] = "BidAmount"
        ws[f"I{self.statelog_start_row}"] = "Comment"
        self.next_statelog_row = self.statelog_start_row + 1
        self.wb.save(self.excel_filename)

    def log_state(self, event, manager, player=None, base_bid=None, bid_amount=None, comment=""):
        """Log an event to the StateLog section of the session sheet."""
        ws = self.session_ws
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        manager_id = self.manager_ids.get(manager, "")
        player_id = self.player_ids.get(player, "") if player else ""
        ws.append([
            timestamp, event, manager_id, manager, player_id, player, base_bid, bid_amount, comment
        ])
        self.wb.save(self.excel_filename)

    # --- Auction Logic Methods ---

    def select_item(self, item):
        """Select an item to start bidding on."""
        self.toggle_bidding_enabled()
        self.current_item = item
        self.current_bid = item[1]
        self.highest_bidder = None
        self.bid_history = []
        self.update_labels()
        # Log event
        self.log_state(
            event="SelectPlayer",
            manager="",
            player=item[0],
            base_bid=item[1],
            comment="Player selected for auction"
        )

    def place_bid(self, team):
        """Handle a bid placed by a team."""
        if self.bidding_enabled and self.current_item and self.team_money[team] >= int(self.current_bid):
            if self.highest_bidder != team:
                if self.highest_bidder is None:
                    next_bid = self.current_bid
                else:
                    # Determine bid increment based on current bid value
                    if int(self.current_bid) < 50:
                        bid_increment = 0
                    elif int(self.current_bid) < 100:
                        bid_increment = 5
                    elif int(self.current_bid) < 200:
                        bid_increment = 10
                    else:
                        bid_increment = 25

                    next_bid = self.current_bid + bid_increment

                # Check if team has enough money for the next bid
                if self.team_money[team] < next_bid:
                    messagebox.showinfo(
                        "Bid Rejected",
                        f"{team} does not have enough money to place this bid."
                    )
                    return

                # Record bid history for undo functionality
                self.bid_history.append((self.highest_bidder, self.current_bid))
                self.current_bid = next_bid
                self.highest_bidder = team
                self.update_labels()
                # Log bid event
                self.log_state(
                    event="Bid",
                    manager=team,
                    player=self.current_item[0],
                    base_bid=self.current_item[1],
                    bid_amount=self.current_bid,
                    comment=f"Manager {team} bid for player {self.current_item[0]}"
                )

    def undo_last_bid(self):
        """Undo the last bid placed."""
        if self.bid_history:
            previous_bidder, previous_bid = self.bid_history.pop()
            self.current_bid = previous_bid
            self.highest_bidder = previous_bidder
            self.update_labels()
            # Log undo event
            self.log_state(
                event="UndoBid",
                manager=previous_bidder if previous_bidder else "",
                player=self.current_item[0] if self.current_item else "",
                base_bid=self.current_item[1] if self.current_item else "",
                bid_amount=previous_bid,
                comment="Undo last bid"
            )
        else:
            messagebox.showinfo(
                "Undo",
                "No previous bid to undo."
            )

    def end_bidding_round(self):
        """End the current bidding round and assign the item to the highest bidder."""
        if self.bidding_enabled and self.highest_bidder:
            item_name, item_value = self.current_item
            # Deduct bid amount from winning team's money and add item to their inventory
            self.team_money[self.highest_bidder] -= int(self.current_bid)
            self.team_inventory[self.highest_bidder][item_name] = int(self.current_bid)
            self.remove_current_item()
            messagebox.showinfo(
                "Auction Result",
                f"{self.highest_bidder} won {item_name} for ₹{self.current_bid}!"
            )
            self.toggle_bidding_disabled()
            self.update_labels()
            # Log bought event
            self.log_state(
                event="Bought",
                manager=self.highest_bidder,
                player=item_name,
                base_bid=item_value,
                bid_amount=self.current_bid,
                comment=f"Manager {self.highest_bidder} bought player {item_name}"
            )
        else:
            messagebox.showinfo(
                "Auction Result",
                "No bids were placed in this round."
            )

    def remove_current_item(self):
        """Remove the current item from the list after it is won."""
        if self.current_item in self.items:
            self.items.remove(self.current_item)
        self.item_list_text.config(state=tk.NORMAL)
        self.item_list_text.delete("1.0", tk.END)
        for item, price in self.items:
            item_button = tk.Button(
                self,
                text=f"{item} (Starting Price: ₹{price})",
                font=('Arial', 12),
                command=lambda item=(item, price): self.select_item(item)
            )
            item_button.grid(sticky="w")
            self.item_list_text.window_create("end", window=item_button)
            self.item_list_text.insert("end", "\n")
        self.item_list_text.config(state=tk.DISABLED)

    def update_labels(self):
        """Update all labels to reflect the current auction state."""
        if self.current_item:
            self.current_item_label.config(
                text=f"Current Item: {self.current_item[0]} (Starting Price: ₹{self.current_item[1]})"
            )
            if self.highest_bidder:
                self.bid_status_label.config(
                    text=f"Current Highest Bid: ₹{self.current_bid} by {self.highest_bidder}"
                )
            else:
                self.bid_status_label.config(
                    text=f"No bids placed yet"
                )
            for team in self.teams:
                # Update money label for each team
                self.money_labels[team].config(
                    text=f"{team} \nMoney: ₹{self.team_money[team]}"
                )
                # Update the inventory list label
                inventory_text = "\n".join(
                    [f"{idx+1}. {item} (Value: ₹{bid_price})"
                     for idx, (item, bid_price) in enumerate(self.team_inventory[team].items())]
                )
                self.inventory_list_labels[team].config(
                    text=inventory_text,
                    wraplength=200,
                    justify="center",
                    anchor="n"
                )
        else:
            self.current_item_label.config(
                text="No items available for auction"
            )

    def toggle_bidding_enabled(self):
        """Enable bidding for the current round."""
        self.bidding_enabled = 1

    def toggle_bidding_disabled(self):
        """Disable bidding after a round ends."""
        self.bidding_enabled = 0

    def update_end_money_in_excel(self):
        """Update the End Money column for each team in the Excel header."""
        ws = self.session_ws
        for idx, team in enumerate(self.teams):
            ws[f"H{7+idx}"] = self.team_money[team]
        self.wb.save(self.excel_filename)

def main():
    root = tk.Tk()
    root.title("Auction Manager")

    def start_auction(teams_df, players_df):
        for widget in root.winfo_children():
            widget.destroy()
        app = AuctionApp(root, teams_df, players_df)
        app.pack(fill="both", expand=True)

    file_page = FileSelectPage(root, start_auction)
    file_page.pack(fill="both", expand=True)
    root.mainloop()

if __name__ == "__main__":
    main()
